# -*- coding: utf-8 -*-

#! python3

# USAGE
# python3 sheet_cell_inverter.py pathToSpreadsheet
# python3 sheet_cell_inverter.py updatedProduceSales.xlsx
# python3 sheet_cell_inverter.py updatedProduceSales_v2.xlsx

import openpyxl, sys

try:
	from openpyxl.cell import column_index_from_string,get_column_letter
except ImportError:
	from openpyxl.utils import column_index_from_string,get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.CRITICAL)

#####################################
# PARSE COMMAND LINE
#####################################

# get the file name and/or path

sheet_to_process = sys.argv[1] # this should be the filename/file path

#####################################
# END PARSE COMMAND LINE
#####################################

#####################################
# ACCESS WORKBOOK
#####################################

wb = openpyxl.load_workbook(sheet_to_process)
sheet = wb.active # switch to the active sheet, there should only be one

logging.debug('Testing to see that sheet loaded right, giving a value:  ')
logging.debug(sheet['A2'])
logging.debug(sheet['A2'].value)

# find out the max rows and max columns so we can set upper ends for loops

upper_row_max = sheet.max_row
logging.debug('The maximum number of rows in the sheet is %i' % (upper_row_max))

upper_col_max = sheet.max_column
logging.debug('The maximum number of columns in the sheet is %i' % (upper_col_max))

#####################################
# END ACCESS WORKBOOK
#####################################

#####################################
# ANALYZE WORKBOOK DATA
#####################################

# loop through the spreadsheet cells and store the data in a format that allows you to invert

inverted_dict = {} # this dict will be used to store the inverted coordinates and their values

def row_analyzer(values_dict,min_value,max_value):
	
	for rowValue in range(min_value,max_value):
		
		# within this specific rowValue we go through each colValue
		for colValue in range(1,upper_col_max+1): # +1 because we're not starting at 0
			# convert the colValue into a letter coordinate
			column_letter = get_column_letter(colValue)
			# combine the column coordinate and row coordinate
			cell_coordinate = column_letter + str(rowValue)
			# get the cell coordinate's value and push it into the values_list
			cell_value = sheet[cell_coordinate].value

			# now invert the cell coordinate before storing
			invert_column_letter = get_column_letter(rowValue) # the row value becomes the column value so get its letter
			logging.debug('The invert_column_letter is:  %s' % (invert_column_letter) )
			# create the invert_cell_coordinate
			invert_cell_coordinate = invert_column_letter + str(colValue) # the column value becomes the row value
			logging.debug('The invert_cell_coordinate is:  %s' % (invert_cell_coordinate) )

			# store the invert_cell_coordinate as the key, with value as cell_value
			values_dict[cell_coordinate] = cell_value
			# logging.debug('The value for %s has been stored in the values_dict' % (cell_coordinate))
			# logging.debug('The value for %s' % (cell_coordinate) )
			# logging.debug(cell_value)


# +1 because we're not starting at 0
# do not set the upper limit to upper_row_max+1 as normal, instead set it to n + 1 as we will be inserting the gap at that point
# we still add + 1 because range() stops 1 point before position_row_N normally, we want it to stop exactly at position_row_N

row_analyzer(inverted_dict,1,upper_row_max + 1)

logging.debug('The inverted dict is:  ')
logging.debug(inverted_dict)

#####################################
# END ANALYZE WORKBOOK DATA
#####################################

#####################################
# BUILD NEW WORKBOOK
#####################################

def row_builder(values_dict,workbook):

	sheet = workbook.active

	# for testing
	# for k,v in values_dict.items():
	# 	print(k)
	# 	print(v)

	for k,v in values_dict.items():
		logging.debug('The key value to use is:  %s' % (k))
		logging.debug('The value of the key is:  ')
		logging.debug(v)
		sheet[k] = v


# create new spreadsheet to store values

nwb = openpyxl.Workbook()

# write the values from values_uptoN_list

row_builder(inverted_dict,nwb)

# save the final sheet

nwb.save('invertedSheet.xlsx')
logging.debug('Spreadsheet file saved.')


#####################################
# END BUILD NEW WORKBOOK
#####################################



